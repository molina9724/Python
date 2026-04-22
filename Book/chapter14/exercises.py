import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string

excel_file = openpyxl.load_workbook(
    "/Users/daniel_molina/Downloads/Python/Python/Book/chapter14/example3.xlsx"
)
print(type(excel_file))

print(excel_file.sheetnames)
sheet1 = excel_file["Sheet1"]
sheet2 = excel_file["Sheet2"]
sheet3 = excel_file["Sheet3"]

print(type(sheet1))
print(type(sheet2))
print(type(sheet3))

print(sheet1.title)
print(sheet2.title)
print(sheet3.title)
print(sheet1.active_cell)
print(excel_file.active)

# Will automatically interpret the dates in column A and return them as datetime values rather than strings.
print(sheet1["A1"].value)
c: Cell = sheet1["B1"]
print(f"Row {c.row}, Column {c.column} is {c.value}")
print(f"Cell {c.coordinate} is {c.value}")
print(c.value)

print(sheet1.cell(row=1, column=2))
print(sheet1.cell(row=1, column=2).value)

for i in range(1, 8, 2):
    print(i, sheet1.cell(row=i, column=2).value)

print(sheet1.max_column)
print(sheet1.max_row)

print(get_column_letter(20))
print(get_column_letter(50))
print(get_column_letter(1997))

print(get_column_letter(sheet1.max_column))
print(column_index_from_string("XXX"))

print("------------------------------------------------------------------------")

print(sheet1["A1":"C3"])
for row_of_cell in sheet1["A1":"C3"]:
    for cell in row_of_cell:
        print(cell.coordinate, cell.value)
    print("--- END OF ROW ---")

list(sheet1.columns)[1]
for cell in list(sheet1.columns)[1]:
    print(cell.value)

list(sheet1.rows)[1]
for cell in list(sheet1.rows)[1]:
    print(cell.value)
