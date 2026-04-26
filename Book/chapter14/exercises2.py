# ======================================================================
# 📊 OPENPYXL (EXCEL) EXERCISES
# ======================================================================
# Practice exercises - Write everything from scratch!
# ======================================================================

# pip install openpyxl

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, GradientFill
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter, column_index_from_string


# =====================================================================
#                    SECTION 1: READING EXCEL FILES
# =====================================================================


# ----------------------------------------------------------------------
# 🟢 1: OPEN AND READ A WORKBOOK
#
# Learn: load_workbook(), accessing sheets
#
# Tasks:
# 1. Create a sample Excel file manually (or use openpyxl to create one)
# 2. Load the workbook using load_workbook()
# 3. Print all sheet names in the workbook
# 4. Get the active sheet
# 5. Print the title of the active sheet
# ----------------------------------------------------------------------

wb = load_workbook(
    "/Users/daniel_molina/Downloads/Python/Python/Book/chapter14/example3.xlsx"
)
print(type(wb))
print(wb.sheetnames)
print(wb.active)
active_page = wb.active
if active_page:
    print(active_page.title)

# ----------------------------------------------------------------------
# 🟢 2: ACCESS CELLS BY COORDINATE
#
# Learn: sheet['A1'], cell.value
#
# Tasks:
# 1. Load a workbook with data
# 2. Access cell A1 and print its value
# 3. Access cell B2 and print its value
# 4. Access a range like A1:C3
# 5. Print values from multiple specific cells
# ----------------------------------------------------------------------

sheet = wb["Sheet1"]
print(sheet["A1"].value)
print(sheet["B2"].value)
for row in sheet["A1:C3"]:
    for cell in row:
        print(cell.value)

# This storages all the data easily
data = [[cell.value for cell in row] for row in sheet["A1:C3"]]

# ----------------------------------------------------------------------
# 🟢 3: ACCESS CELLS BY ROW AND COLUMN
#
# Learn: sheet.cell(row=, column=)
#
# Tasks:
# 1. Load a workbook
# 2. Access cell at row 1, column 1 using sheet.cell()
# 3. Access cell at row 5, column 3
# 4. Compare with coordinate method (should get same result)
# 5. Print both row and column number of a cell
# ----------------------------------------------------------------------

print(sheet.cell(row=1, column=1).value)
print(sheet.cell(row=5, column=3).value)

print(sheet["A1"].value)
print(sheet["C5"].value)

row = 5
column = get_column_letter(3)
coordinate = column + str(row)
print(coordinate)

print(sheet[coordinate].value)

# ----------------------------------------------------------------------
# 🟡 4: ITERATE OVER ROWS
#
# Learn: sheet.iter_rows(), sheet.rows
#
# Tasks:
# 1. Load a workbook with multiple rows of data
# 2. Use iter_rows() to loop through all rows
# 3. Print each cell value in each row
# 4. Limit iteration to specific range (min_row, max_row, etc.)
# 5. Count total number of rows with data
# ----------------------------------------------------------------------

empty_cells = []
non_empty_cells = []
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
    for cell in row:
        if cell.value is not None:
            non_empty_cells.append(cell.coordinate)
        else:
            empty_cells.append(cell.coordinate)

print(empty_cells)

# ----------------------------------------------------------------------
# 🟡 5: ITERATE OVER COLUMNS
#
# Learn: sheet.iter_cols(), sheet.columns
#
# Tasks:
# 1. Load a workbook with multiple columns
# 2. Use iter_cols() to loop through columns
# 3. Print all values in each column
# 4. Limit to specific column range
# 5. Find the column with the maximum value
# ----------------------------------------------------------------------

import math

maximum_value = -math.inf
maximum_value_cell = None

for column in sheet.iter_cols(min_row=1, max_row=sheet.max_row):
    for cell in column:
        print(cell.value)
        if isinstance(cell.value, (int, float)):
            if cell.value > maximum_value:  # type: ignore
                maximum_value = cell.value
                maximum_value_cell = cell.coordinate

print(maximum_value)
print(maximum_value_cell)

# ----------------------------------------------------------------------
# 🟡 6: GET SHEET DIMENSIONS
#
# Learn: sheet.max_row, sheet.max_column, sheet.dimensions
#
# Tasks:
# 1. Load a workbook
# 2. Print the maximum row number with data
# 3. Print the maximum column number with data
# 4. Print the sheet dimensions as a string
# 5. Calculate total number of cells with data
# ----------------------------------------------------------------------

print(sheet.max_row)
print(sheet.max_column)

print(sheet.dimensions)

cells_with_data = 0
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
    for cell in row:
        cells_with_data += 1

print(cells_with_data)

# =====================================================================
#                    SECTION 2: WRITING TO EXCEL
# =====================================================================


# ----------------------------------------------------------------------
# 🟢 7: CREATE NEW WORKBOOK
#
# Learn: Workbook(), save()
#
# Tasks:
# 1. Create a new Workbook object
# 2. Access the active sheet
# 3. Write a value to cell A1
# 4. Save the workbook to a file
# 5. Open the file manually to verify
# ----------------------------------------------------------------------

workbook = openpyxl.Workbook()
print(workbook.sheetnames)

sheet = workbook.active

sheet.title = "my_sheet"  # type: ignore
sheet["A1"] = "Here I am"  # type: ignore
workbook.save(
    "/Users/daniel_molina/Downloads/Python/Python/Book/chapter14/myExcelFile.xlsx"
)
print("Hello")

file_path = (
    "/Users/daniel_molina/Downloads/Python/Python/Book/chapter14/myExcelFile.xlsx"
)

# ----------------------------------------------------------------------
# 🟢 8: WRITE TO CELLS
#
# Learn: sheet['A1'] = value, sheet.cell().value = value
#
# Tasks:
# 1. Create a new workbook
# 2. Write strings to column A (names)
# 3. Write numbers to column B (ages)
# 4. Write dates to column C
# 5. Save and verify the file
# ----------------------------------------------------------------------

from datetime import datetime
import string

workbook = openpyxl.load_workbook(
    "/Users/daniel_molina/Downloads/Python/Python/Book/chapter14/myExcelFile.xlsx",
    data_only=True,
)
sheet = workbook["my_sheet"]

sheet["A1"] = "Names"
sheet["B1"] = "Ages"
sheet["C1"] = "Dates"


for i in range(2, 12):
    sheet.cell(row=i, column=1).value = string.ascii_lowercase[i - 2]
    # sheet["A"+str(i)] = string.ascii_lowercase[i - 2]
    sheet.cell(row=i, column=2).value = i - 1
    # sheet["B"+str(i)] = i-1
    sheet.cell(row=i, column=3).value = datetime.now()
    # sheet["C"+str(i)] = datetime.now()
workbook.save(file_path)

# ----------------------------------------------------------------------
# 🟡 9: CREATE MULTIPLE SHEETS
#
# Learn: workbook.create_sheet(), sheet positioning
#
# Tasks:
# 1. Create a new workbook
# 2. Create a sheet named "Data" at the end
# 3. Create a sheet named "Summary" at the beginning
# 4. Create a sheet named "Charts" at index 1
# 5. Print all sheet names in order
# 6. Delete one of the sheets
# ----------------------------------------------------------------------

workbook.create_sheet(index=1, title="Data")
workbook.create_sheet(index=0, title="Summary")
workbook.create_sheet(index=1, title="Charts")
workbook.save(file_path)
print(workbook.sheetnames)

# ----------------------------------------------------------------------
# 🟡 10: RENAME AND COPY SHEETS
#
# Learn: sheet.title, workbook.copy_worksheet()
#
# Tasks:
# 1. Load or create a workbook with data
# 2. Rename a sheet
# 3. Copy a sheet to create a duplicate
# 4. Verify both sheets have the same data
# 5. Modify copy without affecting original
# ----------------------------------------------------------------------

print(workbook.active)
sheet.title = "renamed_sheet"
sheet = workbook["renamed_sheet"]
workbook.active = sheet

workbook.copy_worksheet(workbook.active)
workbook.save(file_path)

copy_sheet = workbook["renamed_sheet Copy"]
# copy_sheet["A1"] = "different value" this one triggers the assert

for sheet_row, copy_row in zip(sheet.iter_rows(), copy_sheet.iter_rows()):
    for sheet_cell, copy_row_cell in zip(sheet_row, copy_row):
        assert sheet_cell.value == copy_row_cell.value

copy_sheet["A1"] = "Check how I modify the copy but not the original one"
workbook.save(file_path)

# ----------------------------------------------------------------------
# 🟡 11: WRITE FORMULAS
#
# Learn: Writing Excel formulas as strings
#
# Tasks:
# 1. Create a workbook with numbers in column A and B
# 2. Write SUM formula in column C
# 3. Write AVERAGE formula
# 4. Write IF formula
# 5. Write VLOOKUP or other complex formula
# 6. Save and open in Excel to verify formulas work
# ----------------------------------------------------------------------

sheet["B12"] = "=SUM(B2:B11)"
workbook.save(file_path)

sheet["B13"] = "=AVERAGE(B2:B11)"
workbook.save(file_path)

sheet["B14"] = '=IF(A1="Carlos",True,False)'
workbook.save(file_path)

sheet["B15"] = '=VLOOKUP("c",A1:C11,2,FALSE)'
workbook.save(file_path)

# =====================================================================
#                    SECTION 3: CELL STYLING
# =====================================================================


# ----------------------------------------------------------------------
# 🟡 12: FONT STYLING
#
# Learn: Font object, applying to cells
#
# Tasks:
# 1. Create a workbook with text
# 2. Make header row bold
# 3. Change font size of specific cells
# 4. Change font color
# 5. Apply italic and underline
# 6. Use different font names
# ----------------------------------------------------------------------

my_font = Font(
    name="Calibri",
    size=32,
    bold=True,
    color="3F27F5",
    italic=True,
    underline="single",
)
sheet["A2"].font = my_font

my_alignment = Alignment(
    horizontal="center",
    vertical="center",
)

sheet["A2"].alignment = my_alignment

my_fill = PatternFill(
    fill_type="mediumGray",
    start_color="BBF527",
    end_color="F5275E",
)

sheet["A2"].fill = my_fill

workbook.save(file_path)

# ----------------------------------------------------------------------
# 🟡 13: CELL FILL AND COLORS
#
# Learn: PatternFill object
#
# Tasks:
# 1. Create a workbook with data
# 2. Apply background color to header row
# 3. Apply different colors to alternate rows
# 4. Use gradient fill on a cell
# 5. Create a color-coded legend
# ----------------------------------------------------------------------

my_gradient_fill = GradientFill(stop=("000000", "FFFFFF"))

for cell in sheet[1]:
    cell.font = my_font
    cell.alignment = my_alignment
    cell.fill = my_gradient_fill

my_legend = [
    ("Overdue", "FF0000"),
    ("Due Soon", "FFFF00"),
    ("On Time", "00FF00"),
]

row = 1
for label, color in my_legend:
    sheet[f"A{row}"].fill = PatternFill(start_color=color, fill_type="solid")
    sheet[f"B{row}"].value = label
    row += 1

workbook.save(file_path)

# ----------------------------------------------------------------------
# 🟡 14: BORDERS
#
# Learn: Border, Side objects
#
# Tasks:
# 1. Create a workbook with a data table
# 2. Add thin borders around all cells in the table
# 3. Add thick border around the entire table
# 4. Apply different border styles (dashed, dotted)
# 5. Add borders only to specific sides of cells
# ----------------------------------------------------------------------

my_borders = Border(
    left=Side(border_style="dotted", color="000000"),
    right=Side(border_style="dotted", color="000000"),
    top=Side(border_style="dashed", color="000000"),
    bottom=Side(border_style="dashed", color="000000"),
)

sheet["B15"].border = my_borders

workbook.save(file_path)


# ----------------------------------------------------------------------
# 🟡 15: ALIGNMENT
#
# Learn: Alignment object
#
# Tasks:
# 1. Create a workbook with various data
# 2. Center-align text horizontally
# 3. Center-align text vertically
# 4. Apply text wrapping for long text
# 5. Rotate text at an angle
# 6. Apply different alignments to different columns
# ----------------------------------------------------------------------


# =====================================================================
#                    SECTION 4: ROWS AND COLUMNS
# =====================================================================


# ----------------------------------------------------------------------
# 🟡 16: ADJUST DIMENSIONS
#
# Learn: row_dimensions, column_dimensions
#
# Tasks:
# 1. Create a workbook with data
# 2. Set specific row height
# 3. Set specific column width
# 4. Auto-fit column width based on content (approximate)
# 5. Hide a row
# 6. Hide a column
# ----------------------------------------------------------------------


# ----------------------------------------------------------------------
# 🟡 17: INSERT AND DELETE ROWS
#
# Learn: insert_rows(), delete_rows()
#
# Tasks:
# 1. Create a workbook with data in rows 1-10
# 2. Insert 2 rows at row 5
# 3. Verify data shifted down
# 4. Delete row 3
# 5. Verify data shifted up
# ----------------------------------------------------------------------


# ----------------------------------------------------------------------
# 🟡 18: INSERT AND DELETE COLUMNS
#
# Learn: insert_cols(), delete_cols()
#
# Tasks:
# 1. Create a workbook with data in columns A-E
# 2. Insert a column at position C
# 3. Verify data shifted right
# 4. Delete column B
# 5. Verify data shifted left
# ----------------------------------------------------------------------


# ----------------------------------------------------------------------
# 🟡 19: MERGE AND UNMERGE CELLS
#
# Learn: merge_cells(), unmerge_cells()
#
# Tasks:
# 1. Create a workbook
# 2. Merge cells A1:D1 for a title
# 3. Write text to merged cell
# 4. Center the text in merged cell
# 5. Merge cells for a subtitle row
# 6. Unmerge cells and observe result
# ----------------------------------------------------------------------


# ----------------------------------------------------------------------
# 🟡 20: FREEZE PANES
#
# Learn: sheet.freeze_panes
#
# Tasks:
# 1. Create a workbook with headers in row 1
# 2. Add many rows of data
# 3. Freeze the top row so headers stay visible
# 4. Create another sheet and freeze first column
# 5. Freeze both row 1 and column A simultaneously
# ----------------------------------------------------------------------


# =====================================================================
#                    SECTION 5: CHARTS
# =====================================================================


# ----------------------------------------------------------------------
# 🔴 21: BAR CHART
#
# Learn: BarChart, Reference objects
#
# Tasks:
# 1. Create a workbook with category and value data
# 2. Create a Reference for the data
# 3. Create a Reference for the categories
# 4. Create a BarChart object
# 5. Add data and categories to chart
# 6. Add chart to the sheet
# 7. Set chart title and axis labels
# ----------------------------------------------------------------------


# ----------------------------------------------------------------------
# 🔴 22: LINE CHART
#
# Learn: LineChart object
#
# Tasks:
# 1. Create data suitable for a line chart (e.g., time series)
# 2. Create a LineChart
# 3. Add multiple data series
# 4. Customize line styles
# 5. Add to sheet and save
# ----------------------------------------------------------------------


# ----------------------------------------------------------------------
# 🔴 23: PIE CHART
#
# Learn: PieChart object
#
# Tasks:
# 1. Create data suitable for a pie chart
# 2. Create a PieChart
# 3. Add data and labels
# 4. Add chart to sheet
# 5. Save and verify in Excel
# ----------------------------------------------------------------------


# =====================================================================
#                    SECTION 6: UTILITY FUNCTIONS
# =====================================================================


# ----------------------------------------------------------------------
# 🟢 24: COLUMN LETTER CONVERSIONS
#
# Learn: get_column_letter(), column_index_from_string()
#
# Tasks:
# 1. Convert column number 1 to letter (should be 'A')
# 2. Convert column number 27 to letter (should be 'AA')
# 3. Convert letter 'C' to number (should be 3)
# 4. Convert letter 'AB' to number
# 5. Create a loop that prints letters for columns 1-30
# ----------------------------------------------------------------------


# ----------------------------------------------------------------------
# 🟡 25: WORKING WITH CELL RANGES
#
# Learn: Processing ranges of cells
#
# Tasks:
# 1. Load a workbook with data
# 2. Select a range using sheet['A1:C10']
# 3. Iterate over the range
# 4. Calculate sum of all numeric values in range
# 5. Find minimum and maximum values in range
# ----------------------------------------------------------------------


# =====================================================================
#                    SECTION 7: REAL-WORLD PROJECTS
# =====================================================================


# ----------------------------------------------------------------------
# 🔴 26: GRADE SPREADSHEET
#
# Scenario: Create a student grade tracker
#
# Tasks:
# 1. Create headers: Name, Assignment1, Assignment2, Midterm, Final, Average
# 2. Add data for 10 students
# 3. Write AVERAGE formula for each student
# 4. Apply conditional formatting (manually or via styles):
#    - Green for average >= 90
#    - Yellow for average >= 70
#    - Red for average < 70
# 5. Add a summary section with class average
# 6. Style the spreadsheet professionally
# ----------------------------------------------------------------------


# ----------------------------------------------------------------------
# 🔴 27: SALES REPORT
#
# Scenario: Create a monthly sales report
#
# Tasks:
# 1. Create data: Product, Jan, Feb, Mar, Apr, Total, Percentage
# 2. Add data for multiple products
# 3. Write SUM formulas for totals
# 4. Write percentage formulas
# 5. Add a chart showing sales by product
# 6. Add a chart showing monthly trends
# 7. Format currency columns appropriately
# 8. Add borders and professional styling
# ----------------------------------------------------------------------


# ----------------------------------------------------------------------
# 🔴 28: READ AND UPDATE EXISTING FILE
#
# Scenario: Process an existing spreadsheet
#
# Tasks:
# 1. Create an "input" Excel file with raw data
# 2. Write a script that:
#    - Opens the file
#    - Reads all data
#    - Performs calculations
#    - Adds new columns with results
#    - Saves to a new "output" file
# 3. Do not modify the original file
# ----------------------------------------------------------------------


# ----------------------------------------------------------------------
# 🔴 29: MULTIPLE SHEETS REPORT
#
# Scenario: Create workbook with related sheets
#
# Tasks:
# 1. Sheet 1 "Data": Raw data table
# 2. Sheet 2 "Summary": Aggregated statistics
# 3. Sheet 3 "Charts": Visual representations
# 4. Use formulas that reference other sheets
# 5. Create navigation-friendly layout
# ----------------------------------------------------------------------


# ----------------------------------------------------------------------
# 🔴 30: BULK FILE PROCESSOR
#
# Scenario: Process multiple Excel files
#
# Tasks:
# 1. Create 3 sample Excel files with similar structure
# 2. Write script to:
#    - Find all .xlsx files in a directory
#    - Open each file
#    - Extract specific data from each
#    - Combine into a master spreadsheet
# 3. Add a "Source File" column to track origin
# 4. Save the combined result
# ----------------------------------------------------------------------


# ======================================================================
# 📊 QUICK REFERENCE - Loading and Saving
# ======================================================================
#
# Load existing file:     wb = load_workbook('file.xlsx')
# Create new workbook:    wb = Workbook()
# Save workbook:          wb.save('file.xlsx')
# Get active sheet:       sheet = wb.active
# Get sheet by name:      sheet = wb['SheetName']
# Get all sheet names:    wb.sheetnames
#
# ======================================================================


# ======================================================================
# 📊 QUICK REFERENCE - Accessing Cells
# ======================================================================
#
# By coordinate:          cell = sheet['A1']
# By row/column:          cell = sheet.cell(row=1, column=1)
# Get value:              value = cell.value
# Set value:              sheet['A1'] = 'Hello'
# Get range:              cells = sheet['A1:C3']
#
# ======================================================================


# ======================================================================
# 📊 QUICK REFERENCE - Iteration
# ======================================================================
#
# Iterate rows:           for row in sheet.iter_rows():
# Iterate columns:        for col in sheet.iter_cols():
# With limits:            sheet.iter_rows(min_row=1, max_row=10)
# Values only:            sheet.iter_rows(values_only=True)
# Max dimensions:         sheet.max_row, sheet.max_column
#
# ======================================================================


# ======================================================================
# 📊 QUICK REFERENCE - Styling
# ======================================================================
#
# Font:          Font(bold=True, size=14, color='FF0000')
# Fill:          PatternFill(start_color='FFFF00', fill_type='solid')
# Border:        Border(left=Side(style='thin'))
# Alignment:     Alignment(horizontal='center', vertical='center')
# Apply:         cell.font = Font(...)
#
# ======================================================================


# ======================================================================
# 📊 QUICK REFERENCE - Sheets
# ======================================================================
#
# Create sheet:           wb.create_sheet('Name')
# Create at index:        wb.create_sheet('Name', 0)
# Delete sheet:           del wb['SheetName']
# Rename sheet:           sheet.title = 'NewName'
# Copy sheet:             wb.copy_worksheet(sheet)
#
# ======================================================================


# ======================================================================
# 📊 QUICK REFERENCE - Rows and Columns
# ======================================================================
#
# Insert rows:            sheet.insert_rows(7, 2)  # at row 7, insert 2
# Delete rows:            sheet.delete_rows(3, 1)  # at row 3, delete 1
# Insert columns:         sheet.insert_cols(3)
# Delete columns:         sheet.delete_cols(2)
# Row height:             sheet.row_dimensions[1].height = 30
# Column width:           sheet.column_dimensions['A'].width = 20
# Merge cells:            sheet.merge_cells('A1:D1')
# Freeze panes:           sheet.freeze_panes = 'A2'
#
# ======================================================================


# ======================================================================
# 📊 QUICK REFERENCE - Charts
# ======================================================================
#
# Data reference:    data = Reference(sheet, min_col, min_row, max_col, max_row)
# Category ref:      cats = Reference(sheet, ...)
# Create chart:      chart = BarChart() / LineChart() / PieChart()
# Add data:          chart.add_data(data, titles_from_data=True)
# Set categories:    chart.set_categories(cats)
# Add to sheet:      sheet.add_chart(chart, 'E1')
#
# ======================================================================


# ======================================================================
# 🚀 HOW TO RUN
# ======================================================================
#
# Install:     pip install openpyxl
# Run script:  python your_script.py
#
# ======================================================================
