from openpyxl import Workbook

wb = Workbook()
ws = wb.active

print(ws.title)

n = 5

for i in range(1, n + 1):
    ws.cell(row=i + 1, column=1).value = i
    for j in range(1, n + 1):
        ws.cell(row=1, column=j + 1).value = j
        ws.cell(row=i + 1, column=j + 1).value = i * j


m = 2
n = 3

ws.insert_rows(n, m)

wb.save(
    "/Users/daniel_molina/Downloads/Python/Python/Book/chapter14/multiplication.xlsx"
)
